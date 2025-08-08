from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.db.models import Count
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import zipfile
import os
from django.conf import settings
from django.contrib import messages
from .models import TalentEventRegistration, RegistrationActivity, EventStatistics


@admin.register(TalentEventRegistration)
class TalentEventRegistrationAdmin(admin.ModelAdmin):
    """Admin interface for TalentEventRegistration"""

    list_display = [
        'serial_number', 'full_name', 'gender', 'date_of_birth', 'event', 'age_group',
        'city', 'whatsapp_number', 'photo_preview', 'created_at', 'is_active'
    ]

    list_filter = [
        'event', 'age_group', 'gender', 'city', 'terms', 'is_active', 'created_at'
    ]

    search_fields = [
        'serial_number', 'full_name', 'whatsapp_number', 'city', 'id'
    ]

    ordering = ['serial_number']  # Order by serial number (1, 2, 3, 4...)

    readonly_fields = [
        'id', 'serial_number', 'created_at', 'updated_at',
        'ip_address', 'user_agent', 'photo_size_mb'
    ]

    actions = ['make_active', 'make_inactive',
               'export_to_csv', 'export_to_excel', 'download_photos_zip']

    def save_model(self, request, obj, form, change):
        """Override save to ensure serial number is assigned"""
        # Auto-assign serial number if it's a new object (not editing existing)
        if not change and not obj.serial_number:
            last_registration = TalentEventRegistration.objects.order_by(
                'serial_number').last()
            if last_registration and last_registration.serial_number:
                obj.serial_number = last_registration.serial_number + 1
            else:
                obj.serial_number = 1

        # Set IP address from request
        if not obj.ip_address:
            obj.ip_address = self.get_client_ip(request)

        # Set user agent from request
        if not obj.user_agent:
            obj.user_agent = request.META.get('HTTP_USER_AGENT', '')

        super().save_model(request, obj, form, change)

    def get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

    fieldsets = (
        ('Serial & ID', {
            'fields': ('serial_number', 'id')
        }),
        ('Personal Information', {
            'fields': ('full_name', 'gender', 'date_of_birth', 'age_group')
        }),
        ('Event Information', {
            'fields': ('event', 'city')
        }),
        ('Contact Information', {
            'fields': ('whatsapp_number',)
        }),
        ('Photo Upload', {
            'fields': ('photo', 'photo_size_mb')
        }),
        ('Terms & Conditions', {
            'fields': ('terms',)
        }),
        ('System Information', {
            'fields': ('created_at', 'updated_at', 'is_active', 'ip_address', 'user_agent'),
            'classes': ('collapse',)
        }),
    )

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related()

    def photo_preview(self, obj):
        """Display photo preview in admin"""
        if obj.photo:
            return format_html(
                '<img src="{}" style="width: 50px; height: 50px; object-fit: cover; border-radius: 5px;" />',
                obj.photo.url
            )
        return "No Photo"
    photo_preview.short_description = "Photo Preview"

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related()

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related()

    def make_active(self, request, queryset):
        """Mark registrations as active"""
        updated = queryset.update(is_active=True)
        self.message_user(
            request, f'{updated} registrations marked as active.')
    make_active.short_description = "Mark selected registrations as active"

    def make_inactive(self, request, queryset):
        """Mark registrations as inactive"""
        updated = queryset.update(is_active=False)
        self.message_user(
            request, f'{updated} registrations marked as inactive.')
    make_inactive.short_description = "Mark selected registrations as inactive"

    def export_to_excel(self, request, queryset):
        """Export selected registrations to Excel with serial number ordering"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="talent_registrations.xlsx"'

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Talent Event Registrations"

        # Define headers
        headers = [
            'Serial No.', 'Registration ID', 'Full Name', 'Gender', 'Date of Birth',
            'Age Group', 'Event', 'City', 'WhatsApp Number', 'Terms Agreed',
            'Registration Date', 'Photo Size (MB)', 'Active Status'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Order queryset by serial number (1, 2, 3...)
        ordered_queryset = queryset.order_by('serial_number')

        # Write data rows
        for row_num, registration in enumerate(ordered_queryset, 2):
            data = [
                registration.serial_number,
                registration.registration_id,
                registration.full_name,
                registration.get_gender_display(),
                registration.date_of_birth,
                registration.get_age_group_display(),
                registration.get_event_display(),
                registration.city,
                registration.whatsapp_number,
                registration.get_terms_display(),
                registration.created_at.strftime('%Y-%m-%d %H:%M'),
                registration.photo_size_mb,
                'Active' if registration.is_active else 'Inactive'
            ]

            for col_num, value in enumerate(data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(response)
        return response

    export_to_excel.short_description = "Export selected to Excel (Serial order)"

    def export_to_csv(self, request, queryset):
        """Export selected registrations to CSV with serial number ordering"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="talent_registrations.csv"'

        writer = csv.writer(response)

        # Write headers
        writer.writerow([
            'Serial No.', 'Registration ID', 'Full Name', 'Gender', 'Date of Birth',
            'Age Group', 'Event', 'City', 'WhatsApp Number', 'Terms Agreed',
            'Registration Date', 'Photo Size (MB)', 'Active Status'
        ])

        # Order queryset by serial number (1, 2, 3...)
        ordered_queryset = queryset.order_by('serial_number')

        # Write data rows
        for registration in ordered_queryset:
            writer.writerow([
                registration.serial_number,
                registration.registration_id,
                registration.full_name,
                registration.get_gender_display(),
                registration.date_of_birth,
                registration.get_age_group_display(),
                registration.get_event_display(),
                registration.city,
                registration.whatsapp_number,
                registration.get_terms_display(),
                registration.created_at.strftime('%Y-%m-%d %H:%M'),
                registration.photo_size_mb,
                'Active' if registration.is_active else 'Inactive'
            ])

        return response

    export_to_csv.short_description = "Export selected to CSV (Serial order)"

    def download_photos_zip(self, request, queryset):
        """Download photos of selected registrations as a ZIP file"""
        import tempfile
        from datetime import datetime

        # Filter queryset to only include registrations with photos
        queryset_with_photos = queryset.filter(
            photo__isnull=False).exclude(photo='')

        if not queryset_with_photos.exists():
            messages.warning(
                request, 'No photos found for the selected registrations.')
            return

        # Create a temporary file for the ZIP
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')

        try:
            with zipfile.ZipFile(temp_zip.name, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                photos_count = 0

                # Order by serial number for consistent naming
                ordered_queryset = queryset_with_photos.order_by(
                    'serial_number')

                for registration in ordered_queryset:
                    if registration.photo and os.path.exists(registration.photo.path):
                        # Get file extension
                        _, ext = os.path.splitext(registration.photo.name)

                        # Create a meaningful filename: SerialNo_FullName_Event.ext
                        safe_name = "".join(c for c in registration.full_name if c.isalnum() or c in (
                            ' ', '-', '_')).rstrip()
                        safe_event = "".join(c for c in registration.get_event_display(
                        ) if c.isalnum() or c in (' ', '-', '_')).rstrip()

                        filename = f"{registration.serial_number:03d}_{safe_name}_{safe_event}{ext}"
                        filename = filename.replace(' ', '_')

                        # Add file to ZIP
                        zip_file.write(registration.photo.path, filename)
                        photos_count += 1

                # Add a summary text file
                summary_content = f"""Talent Event Registration Photos Summary
=====================================

Total Photos: {photos_count}
Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Export by: {request.user.username if request.user.is_authenticated else 'Unknown'}

File Naming Convention:
SerialNumber_FullName_EventCategory.extension

Registrations Included:
"""

                for registration in ordered_queryset:
                    if registration.photo and os.path.exists(registration.photo.path):
                        summary_content += f"- {registration.serial_number:03d}: {registration.full_name} ({registration.get_event_display()})\n"

                zip_file.writestr('README.txt', summary_content)

            # Prepare response
            with open(temp_zip.name, 'rb') as zip_data:
                response = HttpResponse(
                    zip_data.read(), content_type='application/zip')
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                response[
                    'Content-Disposition'] = f'attachment; filename="talent_event_photos_{timestamp}.zip"'

            messages.success(
                request, f'Successfully downloaded {photos_count} photos in ZIP file.')
            return response

        except Exception as e:
            messages.error(request, f'Error creating ZIP file: {str(e)}')

        finally:
            # Clean up temporary file
            try:
                os.unlink(temp_zip.name)
            except:
                pass

    download_photos_zip.short_description = "Download photos as ZIP file"


@admin.register(RegistrationActivity)
class RegistrationActivityAdmin(admin.ModelAdmin):
    """Admin interface for RegistrationActivity"""

    list_display = [
        'registration', 'activity_type', 'description', 'timestamp'
    ]

    list_filter = [
        'activity_type', 'timestamp'
    ]

    search_fields = [
        'registration__full_name', 'description'
    ]

    readonly_fields = ['timestamp']

    ordering = ['-timestamp']


@admin.register(EventStatistics)
class EventStatisticsAdmin(admin.ModelAdmin):
    """Admin interface for EventStatistics"""

    list_display = [
        'date', 'total_registrations', 'get_top_event', 'get_top_age_group'
    ]

    readonly_fields = [
        'total_registrations', 'registrations_by_event',
        'registrations_by_age_group', 'registrations_by_city'
    ]

    ordering = ['-date']

    def get_top_event(self, obj):
        """Get the most popular event for the day"""
        if obj.registrations_by_event:
            top_event = max(obj.registrations_by_event,
                            key=obj.registrations_by_event.get)
            return f"{top_event} ({obj.registrations_by_event[top_event]})"
        return "No data"
    get_top_event.short_description = "Top Event"

    def get_top_age_group(self, obj):
        """Get the most popular age group for the day"""
        if obj.registrations_by_age_group:
            top_age = max(obj.registrations_by_age_group,
                          key=obj.registrations_by_age_group.get)
            return f"{top_age} ({obj.registrations_by_age_group[top_age]})"
        return "No data"
    get_top_age_group.short_description = "Top Age Group"


# Customize admin site
admin.site.site_header = "Bhudev Kalakaar 2025 Admin"
admin.site.site_title = "Talent Event Admin"
admin.site.index_title = "Talent Event Management"
